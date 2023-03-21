import numpy as np
import pandas as pd
from pandas import *
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from urllib.request import urlopen,Request


wb = load_workbook(filename= 'dndspeakindex.xlsx')
ws = wb["Sheet1"]

links = []
for x in ws.iter_cols():
    for cell in x:
        links.append(cell.hyperlink.target)
linkdir = {}

for num in links:
    req = Request(url=num, headers={'User-Agent': 'Mozilla/5.0'})
    webpage = urlopen(req).read()
    soup = BeautifulSoup(webpage,'lxml')
    title = soup.body.find('h1').text
    a = []
    for art in soup.find_all('article'):
        a.append(art.find_all('a')[0]['href'])
        linkdir[title] = a

print(linkdir)
"""
with open(r'items.txt', 'w') as fp:
    for item in linkdir:
        # write each item on a new line
        fp.write("%s\n" % item)
        
"""

